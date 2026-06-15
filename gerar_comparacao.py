# -*- coding: utf-8 -*-
"""Gera planilha de comparacao das datas de repactuacao (planilha x contratos SharePoint).
NAO altera a planilha original. Saida: Validacao_Repactuacoes_Locacao.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OK, CONF, DIV, NA = "OK", "CONF", "DIV", "NA"

fill_ok   = PatternFill("solid", fgColor="C6EFCE")
fill_conf = PatternFill("solid", fgColor="FFEB9C")
fill_div  = PatternFill("solid", fgColor="FFC7CE")
fill_na   = PatternFill("solid", fgColor="F2F2F2")
fill_hdr  = PatternFill("solid", fgColor="1F4E78")
fill_sec  = PatternFill("solid", fgColor="D9E1F2")
font_hdr  = Font(color="FFFFFF", bold=True, size=10)
font_sec  = Font(color="1F4E78", bold=True, size=11)
font_b    = Font(bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill_for(st):
    return {OK: fill_ok, CONF: fill_conf, DIV: fill_div, NA: fill_na}.get(st, None)

# Cada row: (cliente, unidade,
#   p_ini, st_ini, c_ini, p_fim, st_fim, c_fim,
#   p_ult, st_ult, c_ult, p_prox, st_prox, c_prox, obs)
sections = [
("BAÍA SUL MEDICAL CENTER", [
 ("Abreu Bastian Serv. de Saúde S/S - ME","Sala 102A-102B",
  "16/06/2015",OK,"15/06/2015 (orig.)","31/03/2031",OK,"31/03/2031",
  "07/04/2026",OK,"5º Adit. 07/04/2026","27/02/2028",OK,"27/02/2028",
  "5º aditivo (07/04/2026) prorrogou a vigência p/ 01/04/2026–31/03/2031. Próxima repactuação a mercado em 27/02/2028 (36 meses do 4º aditivo, 27/02/2025). Datas conferem."),
 ("Serviço de Oncologia Médica (Clínica SOMA)","Sala 202",
  "12/09/2011",CONF,"09/09/2011 (orig.)","12/09/2023",DIV,"Vencido / provável Indeterminado",
  "18/03/2021",OK,"4º Adit. 18/03/2021 (valor)","Valor acima da média",CONF,"—",
  "Último aditivo localizado é o 4º (só valor). Prazo original venceu ~set/2023 sem aditivo de prorrogação localizado — contrato provavelmente em prazo indeterminado. Verificar aditivos 1º–3º e eventual renovação não digitalizada."),
 ("Paulo Pukenis Tubelis (Vision Oftalmologia)","Sala 219",
  "28/04/2006",CONF,"19/04/2006 (orig.)","31/01/2023",DIV,"Indeterminado",
  "24/02/2025",CONF,"6º Adit. 27/02/2025","24/02/2028",OK,"27/02/2028",
  "6º aditivo (27/02/2025) repactuou valor p/ R$7.750 e o contrato segue por prazo INDETERMINADO. Planilha mostra fim antigo (31/01/2023). Última da planilha (24/02) ≈ aditivo (27/02)."),
]),
("CENTRO EMPRESARIAL FLORIANÓPOLIS", [
 ("Carlos Eduardo Witolawski Breda","Sala 113 (G19-20)",
  "01/08/2020",CONF,"15/07/2020","31/07/2025",DIV,"14/07/2025 (vencido; s/ prorrog.)",
  "10/12/2024",OK,"3º Adit. 10/12/2024 (valor)","10/10/2027",DIV,"10/12/2027",
  "3º aditivo (10/12/2024) só reajustou valor. Prazo original venceu em 14/07/2025 sem prorrogação localizada (contrato e 2º aditivo são PDF-imagem). Próxima pela regra: 10/12/2027 (planilha traz 10/10/2027)."),
 ("Bureau Assessoria Empresarial Ltda","Sala 209 (G13)",
  "01/10/2018",CONF,"Não confirmado (PDF ilegível)","30/09/2023",DIV,"Vencido / não confirmado",
  "26/05/2025",CONF,"5º Adit. legível 12/05/2023","25/05/2028",CONF,"26/05/2028 (se últ.=26/05/2025)",
  "Contrato original e 1º aditivo retornaram ilegíveis. Último aditivo legível é o 5º (12/05/2023). A planilha indica última reneg. 26/05/2025 — provável 6º aditivo não localizado. Fim 30/09/2023 está vencido. Verificar documentos."),
 ("CENA / DU LAC (uso próprio)","Sala 213",
  "05/03/2020",NA,"Uso próprio","05/03/2025",NA,"Uso próprio",
  "—",NA,"—","—",NA,"—",
  "Sala de uso próprio da Cena/Du Lac — não há contrato de locação a terceiros. Sem repactuação aplicável."),
 ("Hoepcke Adm., Part. e Empreend. Ltda","Sala 214 (G11-12)",
  "04/12/2023",OK,"04/12/2023","03/12/2028",OK,"03/12/2028",
  "Ainda não pode",OK,"Sem aditivos (contrato 29/11/2023)","01/12/2026",OK,"04/12/2026 (início+3)",
  "Contrato 29/11/2023, vigência 04/12/2023–03/12/2028, sem aditivos. Datas conferem."),
 ("Ieda Sena Souza","Sala 216 (G10)",
  "08/02/2021",OK,"08/02/2021","07/08/2023",DIV,"Indeterminado",
  "12/02/2025",OK,"1º Adit. 12/02/2025","12/02/2028",OK,"12/02/2028",
  "1º aditivo (12/02/2025) repactuou valor e converteu o contrato em prazo INDETERMINADO. Planilha mantém fim antigo (07/08/2023)."),
 ("AGAH Creative House (planilha: AMAX/AGAH)","Sala 221",
  "01/12/2024",OK,"01/12/2024","30/11/2025",CONF,"30/11/2025 (vencido)",
  "—",CONF,"1º Adit. 08/04/2025 (troca locadora)","—",CONF,"—",
  "Locatário atual é AGAH Creative House Ltda (planilha registra 'AMAX/AGAH'). Contrato de 12 meses (01/12/2024–30/11/2025) já vencido; verificar renovação. 1º aditivo só trocou a locadora p/ Du Lac."),
 ("NOMA Sushi Rooftop Rest Ltda","Garagem 72 - Ático",
  "01/06/2025",OK,"01/06/2025","30/05/2027",OK,"30/05/2027",
  "Ainda não pode",OK,"Sem aditivos (contrato 08/07/2025)","01/06/2027",OK,"≈30/05/2027 (fim)",
  "Contrato 24 meses (01/06/2025–30/05/2027), sem aditivos. Próxima ≈ fim do contrato. Confere."),
]),
("CENTRO EXECUTIVO CARL HOEPCKE", [
 ("Santinvest S.A.","BL A (Suden) 9º Pav - S901-907",
  "13/10/2020",OK,"13/10/2020","09/04/2029",OK,"08/04/2029 (3º Adit.)",
  "09/04/2026",OK,"3º Adit. 09/04/2026","09/04/2029",OK,"09/04/2029 (09/04/2026+3)",
  "3º aditivo (09/04/2026) prorrogou por 36 meses e reajustou (valores escalonados). Datas conferem."),
 ("Itaú Unibanco S.A.","BL B (Norden) 2º Pav - S08-09-10",
  "07/03/2023",OK,"07/03/2023","06/03/2028",OK,"06/03/2028",
  "26/05/2026",OK,"1º Adit. 26/05/2026 (valor)","06/03/2028",OK,"= fim (regra daria 26/05/2029)",
  "1º aditivo (26/05/2026) só alterou valor; prazo segue 07/03/2023–06/03/2028. Próxima na planilha = fim do contrato."),
 ("Banco Sofisa S.A","BL B (Norden) 2º Pav - S11",
  "01/06/2023",DIV,"Novo contrato 16/02/2026","31/05/2025",DIV,"A confirmar (novo contrato)",
  "—",CONF,"Contrato 16/02/2026 (PDF ilegível)","15/02/2029",CONF,"A confirmar",
  "Há NOVO contrato do Banco Sofisa (16/02/2026); a planilha ainda traz datas do locatário anterior (Strider: 01/06/2023–31/05/2025). PDF do novo contrato é digitalizado/ilegível — confirmar vigência diretamente."),
 ("Aviva","BL B (Norden) 2º Pav - S12-13",
  "01/10/2025",OK,"01/10/2025","01/10/2027",OK,"30/09/2027",
  "—",OK,"Sem aditivos (contrato 30/09/2025)","30/09/2027",OK,"≈30/09/2027 (fim)",
  "Contrato 30/09/2025, vigência 01/10/2025–30/09/2027, sem aditivos. Confere (≈1 dia no fim)."),
 ("Unicred Central SC/PR","BL B (Norden) 5º Pav - S508-511",
  "13/10/2020",OK,"13/10/2020","12/10/2030",OK,"12/10/2030",
  "18/09/2025",OK,"4º Adit. 18/09/2025 (div. adm.)","01/01/2027",CONF,"01/01/2027 (3a da repact. 01/01/2024)",
  "Vigência 120 meses (até 12/10/2030). O último aditivo (18/09/2025) tratou de divisão administrativa Du Lac/Kalonga (não repactuação de valor). A repactuação de valor a mercado foi o 2º aditivo (efeito 01/01/2024); 3 anos → 01/01/2027 (provável base da planilha). Confirmar critério."),
 ("Apex Partners","BL A1 (Osten) 1º Pav - S101",
  "25/03/2025",OK,"25/03/2025","25/03/2027",OK,"25/03/2027",
  "25/03/2025",OK,"Sem aditivos (início)","25/03/2027",OK,"≈25/03/2027 (fim)",
  "Contrato 24 meses (25/03/2025–25/03/2027), sem aditivos. Próxima = fim. Confere."),
 ("Santolin","BL A1 (Osten) 2º Pav - S201",
  "01/11/2022",OK,"01/11/2022","31/10/2027",OK,"31/10/2027",
  "16/01/2026",OK,"1º Adit. 15/01/2026 (valor)","31/10/2027",CONF,"15/01/2029 (regra) / planilha usa fim",
  "1º aditivo (15/01/2026) repactuou valor (escalonado), sem prorrogar. Planilha colocou próxima = fim do contrato (31/10/2027); pela regra dos 3 anos seria 15/01/2029. Como o contrato encerra antes, confirma-se no fim. Verificar critério."),
 ("Costa & Advogados","BL A1 (Osten) 3º Pav - S301",
  "13/10/2020",OK,"13/10/2020","12/10/2025",CONF,"12/10/2025 (vencido; s/ prorrog.)",
  "23/04/2025",OK,"4º Adit. 28/04/2025 (valor)","23/04/2028",OK,"28/04/2028 (≈)",
  "4º aditivo (28/04/2025) repactuou valor; o prazo original encerrou em 12/10/2025 sem aditivo de prorrogação localizado. Verificar renovação."),
 ("Euqueroinvestir (EQI)","BL A e B (embasamento) - Sl 03",
  "26/04/2021",OK,"26/04/2021","25/04/2026",DIV,"30/04/2029 (2º Adit. 16/03/2026)",
  "16/03/2026",OK,"2º Adit. 16/03/2026 (renov. 3 anos)","30/04/2029",CONF,"16/03/2029 (regra) / planilha usa novo fim",
  "2º aditivo (16/03/2026) renovou por 3 anos (26/04/2026–30/04/2029). A planilha NÃO atualizou o FIM (mantém 25/04/2026), embora já tenha colocado 30/04/2029 na próxima. Corrigir o fim para 30/04/2029."),
 ("A2F Estacionamentos (Maxipark)","Garagens C. Hoepcke - SS2",
  "13/10/2020",CONF,"08/10/2020 (contrato)","31/12/2030",OK,"31/12/2030 (5º Adit.)",
  "—",DIV,"5º Adit. 30/01/2026 (renov. 60m)","—",DIV,"30/01/2029 (30/01/2026+3)",
  "5º aditivo (30/01/2026) renovou por 60 meses (01/01/2026–31/12/2030). Planilha está SEM última e próxima repactuação — preencher 30/01/2026 e 30/01/2029."),
 ("Olivia Cucina Restaurante Ltda","Vagas SS2 - 24A/25A/26A",
  "17/04/2025",OK,"17/04/2025","16/04/2027",OK,"16/04/2027",
  "ainda não pode",OK,"Sem aditivos","16/04/2027",OK,"≈16/04/2027 (fim)",
  "Contrato 24 meses (17/04/2025–16/04/2027), sem aditivos. Confere."),
 ("Marisqueira Sintra (Ribeiro Pereira)","Vaga SS2 - 23A",
  "09/01/2024",OK,"09/01/2024","09/01/2027",OK,"08/01/2027",
  "Ainda não pode",OK,"Sem aditivos","09/01/2027",OK,"09/01/2027 (09/01/2024+3)",
  "Contrato 36 meses (09/01/2024–08/01/2027), sem aditivos. Próxima coincide com o fim. Confere."),
]),
("CENTRO EXECUTIVO FERREIRA LIMA", [
 ("A2F Estacionamentos (Maxipark)","Garagens FL - G01-33 / G34-56 / PVG",
  "01/04/2009",OK,"01/04/2009 (contrato)","31/03/2010",DIV,"Indeterminado (17º Adit.)",
  "13/01/2025",OK,"17º Adit. 13/01/2025","13/01/2028",OK,"13/01/2028 (13/01/2025+3)",
  "Contrato de 2009 com 17 aditivos; o 17º (13/01/2025) repactuou valor e o contrato corre por prazo INDETERMINADO. Planilha mostra fim de 31/03/2010 (muito desatualizado)."),
 ("Zenith Centro de Ensino Odontológico","Loja 01 (G67-71)",
  "20/04/2011 - 20/04/2023",CONF,"20/04/2011 / vig. atual desde 01/01/2023","31/12/2027",OK,"31/12/2027 (9º Adit. 24/08/2022)",
  "01/09/2025",OK,"10º Adit. 30/08/2025 (valor)","01/08/2028",CONF,"30/08/2028 (30/08/2025+3)",
  "Prazo vigente vem do 9º aditivo (24/08/2022): 01/01/2023–31/12/2027. O 10º aditivo (30/08/2025) só alterou valor. Próxima pela regra: 30/08/2028 (planilha traz 01/08/2028)."),
 ("Orbis Odontologia Integrada Ltda","Loja 02 (G57-76-80-81-82)",
  "20/04/2011 - 20/04/2021",CONF,"20/04/2011 / vig. atual desde 01/01/2023","31/12/2027",OK,"31/12/2027 (8º Adit. 24/08/2022)",
  "01/09/2025",OK,"9º Adit. 30/08/2025 (valor)","01/08/2028",CONF,"30/08/2028 (30/08/2025+3)",
  "Prazo vigente do 8º aditivo (24/08/2022): 01/01/2023–31/12/2027. O 9º aditivo (30/08/2025) só valor. Próxima pela regra: 30/08/2028 (planilha traz 01/08/2028)."),
 ("Intercorp Provedor de Internet Ltda","S101-201",
  "15/06/2018",OK,"15/06/2018","14/06/2022",DIV,"Indeterminado (prazo orig. vencido)",
  "14/11/2025",OK,"4º Adit. 14/11/2025 (valor)","01/11/2028",CONF,"14/11/2028 (14/11/2025+3)",
  "Nenhum dos 4 aditivos prorrogou o prazo; o original venceu em 14/06/2022 e o contrato corre por prazo INDETERMINADO. Planilha mostra fim vencido (14/06/2022)."),
 ("Ney Muller","S102",
  "27/07/2023",OK,"27/07/2023","26/07/2028",OK,"26/07/2028",
  "Em negociação",CONF,"Sem aditivos (contrato 26/07/2023)","—",CONF,"26/07/2026 (contrato+3)",
  "Contrato 60 meses (27/07/2023–26/07/2028), sem aditivos. Próxima repactuação prevista p/ 26/07/2026 (3 anos). Planilha marca 'Em negociação' e próxima em branco."),
 ("Lifeshub Group AS","S202",
  "01/05/2024",OK,"01/05/2024","30/04/2028",OK,"30/04/2028",
  "Ainda não pode",OK,"Sem aditivos (contrato 26/04/2024)","01/05/2027",OK,"≈26/04/2027",
  "Contrato 48 meses (01/05/2024–30/04/2028), sem aditivos. Confere."),
 ("Solon Sehn Advogados Associados","S301 (G83)",
  "07/12/2011",OK,"07/12/2011 (sala) / 15/07/2014 (gar.)","31/01/2022",DIV,"Indeterminado (6º Adit.)",
  "18/03/2025",OK,"6º Adit. 18/03/2025","18/03/2028",OK,"18/03/2028",
  "6º aditivo (18/03/2025) repactuou valor; contrato em prazo INDETERMINADO. Planilha mostra fim antigo (31/01/2022)."),
 ("Freitas Advogados","S302 (G90-91-92)",
  "01/03/2020",OK,"01/03/2020","28/02/2030",OK,"28/02/2030",
  "29/04/2025",OK,"6º T. Acordo 29/04/2025","29/04/2028",OK,"29/04/2028",
  "Contrato 120 meses (01/03/2020–28/02/2030). 6º Termo de Acordo (29/04/2025) repactuou valor sem alterar prazo. Datas conferem."),
 ("Capella e Fogaça","S401 (G84)",
  "01/07/2011",OK,"01/07/2011","11/05/2028",OK,"11/05/2028 (10º Adit. 09/05/2023)",
  "01/05/2023",OK,"10º Adit. 09/05/2023","Em negociação",CONF,"09/05/2026 (já no período)",
  "10º aditivo (09/05/2023) prorrogou até 11/05/2028 e repactuou. Próxima repactuação (3 anos) seria ~09/05/2026 — já em curso; 'Em negociação' é coerente."),
 ("MS Consultoria e Assessoria","S402",
  "10/09/2021",OK,"10/09/2021","09/09/2026",DIV,"09/09/2029 (2º Adit. 03/03/2026)",
  "10/03/2026",OK,"2º Adit. 03/03/2026 (prorrog. 3a)","09/09/2027",DIV,"03/03/2029 (03/03/2026+3)",
  "2º aditivo (03/03/2026) prorrogou o contrato p/ 10/09/2026–09/09/2029 e repactuou. Planilha desatualizada: fim deveria 09/09/2029 e próxima ~03/03/2029 (consta 09/09/2027)."),
 ("B2B Escritórios Compartilhados Ltda","S501",
  "01/01/2025",OK,"01/01/2025","31/10/2029",OK,"31/10/2029",
  "—",OK,"Sem aditivos","01/01/2028",OK,"01/01/2028 (início+3)",
  "Contrato 03/02/2025 (instrumento 01/01/2025), vigência 01/01/2025–31/10/2029 (abrange salas 501 e 502). Sem aditivos. Confere."),
 ("B2B Escritórios Compartilhados Ltda","S502",
  "01/01/2025",OK,"01/01/2025","31/10/2029",OK,"31/10/2029",
  "—",OK,"Sem aditivos","01/01/2028",OK,"01/01/2028 (início+3)",
  "Mesmo contrato da sala 501 (01/01/2025–31/10/2029). Confere."),
]),
("PLAZA DANÚBIO RESIDENCE", [
 ("Carlos André Bastian","AP 102 (G80/80A, 97/97A, HB05)",
  "08/11/2021",OK,"08/11/2021","07/05/2024",DIV,"Indeterminado (orig. venceu 07/05/2024)",
  "10/04/2024",OK,"2º Adit. 10/04/2024","10/04/2027",OK,"10/04/2027 (10/04/2024+3)",
  "Contrato 30 meses (08/11/2021–07/05/2024); 2 aditivos (último 10/04/2024) só de valor/locatário, sem prorrogação — contrato em prazo indeterminado. Planilha mostra fim vencido (07/05/2024)."),
]),
("GALPÃO CAIS DO PORTO", [
 ("Pet Center Comércio e Participações (Petz)","Cais do Porto",
  "14/12/2021",CONF,"01/11/2021 (1º Adit.)","14/12/2031",CONF,"31/10/2031 (1º Adit.)",
  "Não permite renegociação",OK,"1º Adit. 05/11/2021","Não permite renegociação",OK,"—",
  "Contrato de 10 anos (1º aditivo: vigência 01/11/2021–31/10/2031). Planilha registra 14/12/2021–14/12/2031 (~6 semanas de diferença). Sem repactuação trienal (contrato fixo). Confirmar datas exatas."),
]),
("SHOPPING DEODORO", [
 ("Acessórios/Vestuário Gomes (planilha: Opico MS)","Loja Rua Deodoro 126",
  "01/02/2020",OK,"01/02/2020 (orig.)","31/01/2025",DIV,"28/02/2029 (7º Adit. 20/03/2026)",
  "(em branco)",DIV,"7º Adit. 20/03/2026 (prorrog.)","(em branco)",DIV,"20/03/2029 (20/03/2026+3)",
  "7º aditivo (20/03/2026) prorrogou a vigência p/ 01/03/2026–28/02/2029 e repactuou (R$57.000; R$60.000 a partir de dez/2026). Planilha mostra fim vencido (31/01/2025) e sem última/próxima — atualizar."),
]),
("TERRENO TROMPOWSKY", [
 ("Fórmula Estacionamento Ltda","Terreno Área A3 - Trompowsky",
  "01/10/2024",OK,"01/10/2024","30/09/2025",DIV,"30/04/2026 (2º Adit. 25/09/2025)",
  "25/09/2025",OK,"2º Adit. 25/09/2025 (prorrog.)","30/04/2026",CONF,"30/04/2026 (= fim renovado; já vencido)",
  "2º aditivo (25/09/2025) prorrogou até 30/04/2026. Planilha manteve fim antigo (30/09/2025), mas colocou 30/04/2026 na próxima. ATENÇÃO: vigência terminou em 30/04/2026 — verificar nova renovação."),
]),
("CASA CRISPIM MIRA", [
 ("Bruno Spindola Conzatti (Mais Laser)","Casa 11",
  "01/06/2026",OK,"01/06/2026","31/05/2029",OK,"31/05/2029",
  "ainda não pode",OK,"Sem aditivos","31/03/2029",CONF,"01/06/2029 (01/06/2026+3)",
  "Contrato 36 meses (01/06/2026–31/05/2029), sem aditivos. Próxima pela regra: 01/06/2029 (≈ fim); planilha traz 31/03/2029."),
]),
("SQUARE CORPORATE", [
 ("Resulta Corretora de Seguros Ltda","Sala 330",
  "(em branco)",DIV,"26/03/2026","(em branco)",DIV,"26/03/2027",
  "(em branco)",CONF,"Sem aditivos","(em branco)",DIV,"26/03/2029 (init+3) ou fim 26/03/2027",
  "Contrato 12 meses (26/03/2026–26/03/2027), sem aditivos. Planilha está SEM datas nesta linha — preencher início 26/03/2026 e fim 26/03/2027."),
]),
("EDIFÍCIO COMERCIAL PRIME TOWER", [
 ("RBV Incorporações Imobiliárias","Sala 401 (G02, G10)",
  "16/10/2025",OK,"16/10/2025","15/10/2030",OK,"15/10/2030",
  "—",OK,"1º Adit. 13/05/2026 (sub-rogação)","16/10/2028",OK,"16/10/2028 (início+3)",
  "Vigência 16/10/2025–15/10/2030. 1º aditivo (13/05/2026) apenas sub-rogou a locadora p/ a Cena (sem mudar prazo/valor). Próxima = 16/10/2028. Confere."),
]),
("MARINE", [
 ("Igor Dantas Almeida","Apto 202 Mikonos (Gar 25-67-68)",
  "24/07/2025",OK,"24/07/2025","23/07/2026",DIV,"DISTRATADO em 22/05/2026",
  "ainda não pode",DIV,"Distrato 22/05/2026","23/07/2026",DIV,"—",
  "Contrato (24/07/2025–23/07/2026) foi RESCINDIDO antecipadamente por Distrato em 22/05/2026. Planilha ainda o mostra como ativo — atualizar situação (imóvel possivelmente vago/relocado)."),
 ("José Rafael Fernandez Diaz","Apto 304 Creta (Gar 13-44-45, HB12)",
  "01/09/2024",OK,"01/09/2024","31/01/2025",DIV,"28/02/2027",
  "ainda não pode",OK,"Sem aditivos","28/02/2027",OK,"28/02/2027 (= fim)",
  "Contrato 30 meses (01/09/2024–28/02/2027), sem aditivos. Planilha mostra FIM errado (31/01/2025); o correto é 28/02/2027 (valor que, inclusive, já consta na coluna 'próxima')."),
]),
("SAINT-TROPEZ", [
 ("Annalisa Blando Dal Zotto","Apto 401 (Gar 16-17-18)",
  "10/07/2024",OK,"10/07/2024","10/01/2027",DIV,"10/01/2028",
  "ainda não pode",OK,"Sem aditivos","10/01/2027",CONF,"10/01/2028 (fim) / regra 10/07/2027",
  "Contrato 42 meses (10/07/2024–10/01/2028). Planilha mostra fim 10/01/2027, que é a data de SAÍDA ANTECIPADA sem multa; o término contratual é 10/01/2028. Ajustar."),
]),
("CÔTE D'AZUR VILLE", [
 ("Imóvel transferido (Ana Ruth)","AP 202 (G70-71, HB01)",
  "—",NA,"—","—",NA,"—",
  "—",NA,"—","—",NA,"—",
  "Planilha indica 'imóvel passado para a Ana Ruth' (transferido). Sem contrato de locação ativo a validar."),
]),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparação Repactuações"

# Titulo
ws.merge_cells("A1:L1")
ws["A1"] = "VALIDAÇÃO DE REPACTUAÇÕES — Planilha de Locação (Consolidado) × Contratos/Aditivos (SharePoint)"
ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
ws.merge_cells("A2:L2")
ws["A2"] = "Base: planilha 'Controle Contratos de Locação 2023-2026' (aba Consolidado) | Data de referência: 15/06/2026 | A planilha original NÃO foi alterada."
ws["A2"].font = Font(italic=True, size=9, color="808080")
ws.merge_cells("A3:L3")
ws["A3"] = "Legenda:  VERDE = confere   |   AMARELO = conferir / divergência leve   |   VERMELHO = divergente (corrigir)"
ws["A3"].font = Font(bold=True, size=9)

headers = ["Cliente","Unidade","Início (Planilha)","Início (Contrato)",
           "Fim (Planilha)","Fim (Contrato/Atual)","Última Reneg. (Planilha)",
           "Último Aditivo/Repact. (Contrato)","Próxima Reneg. (Planilha)",
           "Próxima Repact. (Correta)","Situação","Observações"]
hr = 5
for c,h in enumerate(headers,1):
    cell = ws.cell(hr,c,h); cell.fill=fill_hdr; cell.font=font_hdr
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border

r = hr+1
def overall(sts):
    if DIV in sts: return ("DIVERGENTE", fill_div)
    if CONF in sts: return ("CONFERIR", fill_conf)
    if all(s==NA for s in sts): return ("N/A", fill_na)
    return ("OK", fill_ok)

for sec_name, rows in sections:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
    sc = ws.cell(r,1,sec_name); sc.fill=fill_sec; sc.font=font_sec
    sc.alignment=Alignment(horizontal="left",vertical="center"); sc.border=border
    r+=1
    for row in rows:
        (cli,uni,p_ini,st_ini,c_ini,p_fim,st_fim,c_fim,
         p_ult,st_ult,c_ult,p_prox,st_prox,c_prox,obs)=row
        vals=[cli,uni,p_ini,c_ini,p_fim,c_fim,p_ult,c_ult,p_prox,c_prox,"",obs]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.border=border
            cell.alignment=Alignment(vertical="top",wrap_text=True,
                horizontal="center" if 3<=c<=10 else "left")
        # colore celulas da planilha conforme status
        if fill_for(st_ini): ws.cell(r,3).fill=fill_for(st_ini)
        if fill_for(st_fim): ws.cell(r,5).fill=fill_for(st_fim)
        if fill_for(st_ult): ws.cell(r,7).fill=fill_for(st_ult)
        if fill_for(st_prox): ws.cell(r,9).fill=fill_for(st_prox)
        sit,sit_fill=overall([st_ini,st_fim,st_ult,st_prox])
        sc=ws.cell(r,11,sit); sc.fill=sit_fill; sc.font=font_b
        sc.alignment=Alignment(horizontal="center",vertical="center")
        r+=1

widths=[30,26,16,22,16,24,18,24,18,22,13,55]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A6"

# ---- Aba Resumo ----
ws2=wb.create_sheet("Resumo")
ws2.merge_cells("A1:C1")
ws2["A1"]="RESUMO — Principais divergências a corrigir na planilha original"
ws2["A1"].font=Font(bold=True,size=12,color="1F4E78")
res_hdr=["#","Item / Unidade","O que corrigir"]
for c,h in enumerate(res_hdr,1):
    cell=ws2.cell(3,c,h); cell.fill=fill_hdr; cell.font=font_hdr; cell.border=border
divs=[
 ("MS Consultoria – S402 (Ferreira Lima)","Fim 09/09/2026→09/09/2029 e próxima 09/09/2027→~03/03/2029 (2º aditivo de 03/03/2026 prorrogou 3 anos)."),
 ("Euqueroinvestir – Embasamento (Carl Hoepcke)","Fim 25/04/2026→30/04/2029 (2º aditivo de 16/03/2026 renovou por 3 anos)."),
 ("Shopping Deodoro","Fim 31/01/2025→28/02/2029 e preencher última (20/03/2026) e próxima (20/03/2029) — 7º aditivo de 20/03/2026."),
 ("Terreno Trompowsky","Fim 30/09/2025→30/04/2026 (2º aditivo); vigência já vencida em 30/04/2026 — verificar nova renovação."),
 ("Maxipark – Ferreira Lima","Fim 31/03/2010→'Indeterminado' (17º aditivo de 13/01/2025)."),
 ("Maxipark – Carl Hoepcke","Preencher última (30/01/2026) e próxima (30/01/2029) — 5º aditivo renovou até 31/12/2030."),
 ("Ieda Sena – S216 (Centro Emp. Floripa)","Fim 07/08/2023→'Indeterminado' (1º aditivo de 12/02/2025)."),
 ("Solon Sehn – S301 (Ferreira Lima)","Fim 31/01/2022→'Indeterminado' (6º aditivo de 18/03/2025)."),
 ("Intercorp – S101/201 (Ferreira Lima)","Fim 14/06/2022→'Indeterminado' (prazo original vencido; nenhum aditivo prorrogou)."),
 ("Vision Oftalmologia – S219 (Baía Sul)","Fim 31/01/2023→'Indeterminado' (6º aditivo de 27/02/2025)."),
 ("Clínica SOMA – S202 (Baía Sul)","Fim 12/09/2023 vencido — verificar renovação; provável prazo indeterminado."),
 ("Carlos Breda – S113 (Centro Emp. Floripa)","Fim 31/07/2025 vencido (vig. 14/07/2025, sem prorrogação localizada); próxima 10/10/2027→10/12/2027."),
 ("Plaza Danúbio – AP102","Fim 07/05/2024 vencido→'Indeterminado' (sem aditivo de prorrogação)."),
 ("Costa & Advogados – S301 (Carl Hoepcke)","Fim 12/10/2025 vencido — verificar renovação."),
 ("AGAH/AMAX – S221 (Centro Emp. Floripa)","Contrato (até 30/11/2025) vencido — verificar renovação; ajustar nome p/ AGAH Creative House."),
 ("Marine – Apto 202 (Igor Dantas)","DISTRATADO em 22/05/2026 — remover/atualizar como vago (planilha mostra ativo)."),
 ("Marine – Apto 304 (José Rafael)","Fim 31/01/2025→28/02/2027 (vigência real do contrato)."),
 ("Saint-Tropez – Apto 401 (Annalisa)","Fim 10/01/2027→10/01/2028 (10/01/2027 é só a data de saída antecipada sem multa)."),
 ("Banco Sofisa – S11 (Carl Hoepcke)","Datas são do locatário anterior (Strider); há novo contrato de 16/02/2026 — atualizar (PDF ilegível, confirmar vigência)."),
 ("Square Corporate – S330 (Resulta)","Linha sem datas — preencher início 26/03/2026 e fim 26/03/2027."),
 ("Bureau – S209 (Centro Emp. Floripa)","Fim 30/09/2023 vencido; confirmar última repactuação 26/05/2025 (aditivo não localizado)."),
 ("Galpão Cais do Porto (Petz)","Conferir início/fim: planilha 14/12/2021–14/12/2031 × aditivo 01/11/2021–31/10/2031."),
 ("Zenith / Orbis – Lojas 01 e 02 (Ferreira Lima)","Próxima 01/08/2028→30/08/2028 (3 anos do 9º/10º aditivo de 30/08/2025)."),
]
rr=4
for i,(item,fix) in enumerate(divs,1):
    ws2.cell(rr,1,i).border=border
    a=ws2.cell(rr,2,item); a.border=border; a.alignment=Alignment(wrap_text=True,vertical="top"); a.font=font_b
    b=ws2.cell(rr,3,fix); b.border=border; b.alignment=Alignment(wrap_text=True,vertical="top")
    rr+=1
ws2.column_dimensions["A"].width=5
ws2.column_dimensions["B"].width=42
ws2.column_dimensions["C"].width=85

out="/home/user/Agenda/Validacao_Repactuacoes_Locacao.xlsx"
wb.save(out)
print("Salvo:",out)
print("Linhas de dados:",sum(len(r) for _,r in sections))
